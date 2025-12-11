
from openpyxl import load_workbook
import requests
from io import BytesIO


def load_products_sizes() -> list[tuple[str|int]]:
    """
    Load products sizes from google sheet
    """
    products_sizes = load_workbook(
        filename=BytesIO(
            requests.get("https://docs.google.com/spreadsheets/d/1mv6epUufD3Q0UTisHDygZ7qRas8w9fHVIVh2dchitdo/export?format=xlsx").content
        )
    )
    products_sizes=products_sizes[products_sizes.sheetnames[0]]

    products_sizes=[
        (
            products_sizes[f"A{row}"].value.replace(" ", ""),
            products_sizes[f"B{row}"].value,
            float(products_sizes[f"C{row}"].value),
            float(products_sizes[f"D{row}"].value),
            float(products_sizes[f"E{row}"].value) if products_sizes[f"E{row}"].value!="no" else products_sizes[f"E{row}"].value,
            float(products_sizes[f"F{row}"].value) if products_sizes[f"F{row}"].value!="no" else products_sizes[f"F{row}"].value,
            products_sizes[f"G{row}"].value,
            products_sizes[f"H{row}"].value,
            products_sizes[f"I{row}"].value
        )
        for row in range(1, products_sizes.max_row+1)
        if products_sizes[f"A{row}"].value is not None
    ]

    return products_sizes